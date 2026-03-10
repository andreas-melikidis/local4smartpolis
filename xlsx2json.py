"""
xlsx2json.py  v3  —  SMARTPOLIS
================================
Μετατρέπει τα ΕΛΣΤΑΤ Excel σε ένα municipality-centric JSON.

Χρήση:
    python xlsx2json.py --root "C:\\Users\\andre\\OneDrive\\Desktop\\ΕΛΣΤΑΤ_Excels" --output smartpolis.json

Δομή output:
    {
      "metadata": { ... },
      "data": {
        "municipalities": { "<geo_code>": { name, geo_level, population, buildings, education } },
        "regions":        { "<geo_code>": { name, population, gdp, enterprises } },
        "national":       { employment_unemployment, gdp_gva }
      }
    }

Fixes από v2:
  - .xls υποστήριξη μέσω LibreOffice (xlrd δεν χρειάζεται)
  - Dedicated parsers για ΚΑΘΕ πίνακα (A01/A02/A04/A06/A07, B01/B02/B03, B09/B14, T5/T7, SEL45, SJO02)
  - B03 disambiguation (SAM04 vs SAM05 έχουν ίδιο όνομα, διαφορετική δομή)
  - find_data_start χειρίζεται int ΚΑΙ string geo_level (ΕΛΣΤΑΤ inconsistency)
  - Δομή JSON οργανωμένη ανά δήμο — όχι ανά κατηγορία αρχείου
  - Fields.txt χρησιμοποιείται ως metadata ανά section
"""

import os
import re
import json
import argparse
import tempfile
import subprocess
import shutil
import openpyxl


# ══════════════════════════════════════════════════════════
#  Utilities
# ══════════════════════════════════════════════════════════

def clean(val):
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        return v if v else None
    return val


def to_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("*", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def to_int(val):
    if val is None:
        return None
    if isinstance(val, float):
        return int(val)
    if isinstance(val, int):
        return val
    s = str(val).strip().replace("*", "").replace(".0", "")
    try:
        return int(s)
    except ValueError:
        return None


def excel_files_in(folder):
    return sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~")
    )


def load_fields(folder_path):
    p = os.path.join(folder_path, "Fields.txt")
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return [line.strip() for line in f if line.strip()]
    return []


def find_data_start(rows):
    """
    Βρίσκει την πρώτη γραμμή όπου col[0] είναι geo_level (αριθμός ή αριθμητικό string).
    Τα ΕΛΣΤΑΤ αρχεία μερικές φορές βάζουν int, μερικές string '0', '1', κτλ.
    """
    for i, row in enumerate(rows):
        if row[0] is not None and row[1] is not None:
            v = row[0]
            if isinstance(v, (int, float)):
                return i
            if isinstance(v, str) and re.match(r"^\d+$", v.strip()):
                return i
    return None


def xls_to_xlsx(xls_path):
    """
    Μετατρέπει .xls σε .xlsx χρησιμοποιώντας LibreOffice headless.
    Επιστρέφει path του .xlsx ή None αν αποτύχει.
    """
    libreoffice = shutil.which("libreoffice") or shutil.which("soffice")
    if not libreoffice:
        print("  ⚠️  LibreOffice δεν βρέθηκε — αδυναμία ανάγνωσης .xls")
        return None

    tmpdir = tempfile.mkdtemp()
    try:
        result = subprocess.run(
            [libreoffice, "--headless", "--convert-to", "xlsx", "--outdir", tmpdir, xls_path],
            capture_output=True, timeout=60
        )
        if result.returncode != 0:
            print(f"  ⚠️  LibreOffice error: {result.stderr.decode()[:200]}")
            return None
        # Βρίσκουμε το παραγόμενο αρχείο
        basename = os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx"
        out = os.path.join(tmpdir, basename)
        return out if os.path.exists(out) else None
    except Exception as e:
        print(f"  ⚠️  LibreOffice exception: {e}")
        return None


# ══════════════════════════════════════════════════════════
#  PARSERS — SAM03 (Regions)
# ══════════════════════════════════════════════════════════

def parse_SAM03_A01(ws):
    """Πληθυσμός κατά φύλο & ομάδες ηλικιών — Περιφέρειες."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":           to_int(row[0]),
            "geo_code":            clean(row[1]),
            "name_el":             clean(row[2]),
            "name_en":             clean(row[6]) if len(row) > 6 else None,
            "population_total":    to_int(row[3]),
            "population_males":    to_int(row[4]),
            "population_females":  to_int(row[5]),
        })
    return [r for r in records if r["name_el"]]


def parse_SAM03_A02(ws):
    """Πληθυσμός κατά φύλο & ιθαγένεια — Περιφέρειες."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":           to_int(row[0]),
            "geo_code":            clean(row[1]),
            "country_code":        clean(row[2]),
            "citizenship_el":      clean(row[3]),
            "citizenship_en":      clean(row[7]) if len(row) > 7 else None,
            "population_total":    to_int(row[4]),
            "population_males":    to_int(row[5]),
            "population_females":  to_int(row[6]),
        })
    return [r for r in records if r["citizenship_el"]]


def parse_SAM03_A04(ws):
    """Πληθυσμός κατά ομάδες ηλικιών & τόπο γέννησης — Περιφέρειες."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":        to_int(row[0]),
            "geo_code":         clean(row[1]),
            "name_el":          clean(row[2]),
            "name_en":          clean(row[6]) if len(row) > 6 else None,
            "population_total": to_int(row[3]),
            "born_in_greece":   to_int(row[4]),
            "born_abroad":      to_int(row[5]),
        })
    return [r for r in records if r["name_el"]]


def parse_SAM03_A06(ws):
    """Ιδιωτικά νοικοκυριά & αριθμός μελών κατά μέγεθος νοικοκυριού — Περιφέρειες."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":            to_int(row[0]),
            "geo_code":             clean(row[1]),
            "household_size":       clean(row[2]),
            "household_size_en":    clean(row[5]) if len(row) > 5 else None,
            "num_households":       to_int(row[3]),
            "num_members":          to_int(row[4]),
        })
    return [r for r in records if r["household_size"]]


def parse_SAM03_A07(ws):
    """Ιδιωτικά μονομελή νοικοκυριά κατά φύλο & ηλικία — Περιφέρειες."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":                        to_int(row[0]),
            "geo_code":                         clean(row[1]),
            "name_el":                          clean(row[2]),
            "name_en":                          clean(row[6]) if len(row) > 6 else None,
            "one_person_households_total":       to_int(row[3]),
            "one_person_households_males":       to_int(row[4]),
            "one_person_households_females":     to_int(row[5]),
        })
    return [r for r in records if r["name_el"]]


# ══════════════════════════════════════════════════════════
#  PARSERS — SAM05 (Municipalities — Κτίρια/Κατοικίες)
# ══════════════════════════════════════════════════════════

def parse_SAM05_B01(ws):
    """Κατοικούμενες κατοικίες κατά πυκνότητα κατοικήσεως & κυριότητα — Δήμοι."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":                    to_int(row[0]),
            "geo_code":                     clean(row[1]),
            "occupancy_density":            clean(row[2]),   # τ.μ. ανά μέλος
            "dwellings_total":              to_int(row[3]),
            "dwellings_owner_occupied":     to_int(row[4]),
            "dwellings_rented_total":       to_int(row[5]),
            "dwellings_rented_furnished":   to_int(row[6]),
            "dwellings_rented_unfurnished": to_int(row[7]),
            "dwellings_other_ownership":    to_int(row[8]) if len(row) > 8 else None,
        })
    return [r for r in records if r["occupancy_density"]]


def parse_SAM05_B02(ws):
    """Κατοικούμενες κατοικίες κατά τύπο κτιρίου & κυριότητα — Δήμοι."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":                    to_int(row[0]),
            "geo_code":                     clean(row[1]),
            "building_type":                clean(row[2]),
            "dwellings_total":              to_int(row[3]),
            "dwellings_owner_occupied":     to_int(row[4]),
            "dwellings_rented_total":       to_int(row[5]),
            "dwellings_rented_furnished":   to_int(row[6]),
            "dwellings_rented_unfurnished": to_int(row[7]),
            "dwellings_other_ownership":    to_int(row[8]) if len(row) > 8 else None,
        })
    return [r for r in records if r["building_type"]]


def parse_SAM05_B03(ws):
    """Νοικοκυριά & μέλη κατά μέγεθος νοικοκυριού — Δήμοι."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":        to_int(row[0]),
            "geo_code":         clean(row[1]),
            "household_size":   clean(row[2]),
            "num_households":   to_int(row[3]),
            "num_members":      to_int(row[4]) if len(row) > 4 else None,
        })
    return [r for r in records if r["household_size"]]


# ══════════════════════════════════════════════════════════
#  PARSERS — SAM04 (Municipalities — Εκπαίδευση)
# ══════════════════════════════════════════════════════════

def parse_SAM04_B03(ws):
    """
    Μόνιμος πληθυσμός κατά επίπεδο εκπαίδευσης & κατάσταση ασχολίας — Δήμοι.
    ΠΡΟΣΟΧΗ: Ίδιο όνομα αρχείου (B03) με το SAM05 B03 αλλά εντελώς διαφορετική δομή!
    SAM04 B03: 11 στήλες, SAM05 B03: 5 στήλες.
    """
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None or len(row) < 7:
            continue
        records.append({
            "geo_level":                        to_int(row[0]),
            "geo_code":                         clean(row[1]),
            "education_level":                  clean(row[2]),
            "population_total":                 to_int(row[3]),
            "economically_active_total":        to_int(row[4]),
            "employed":                         to_int(row[5]),
            "unemployed":                       to_int(row[6]),
            "economically_inactive_total":      to_int(row[7]) if len(row) > 7 else None,
            "students":                         to_int(row[8]) if len(row) > 8 else None,
            "retired":                          to_int(row[9]) if len(row) > 9 else None,
            "other_inactive":                   to_int(row[10]) if len(row) > 10 else None,
        })
    return [r for r in records if r["education_level"]]


def parse_SAM04_B09(ws):
    """Απασχολούμενοι κατά επίπεδο εκπαίδευσης — Δήμοι."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":                        to_int(row[0]),
            "geo_code":                         clean(row[1]),
            "name_el":                          clean(row[2]),
            "employed_total":                   to_int(row[3]),
            "employed_phd_masters_university":  to_int(row[4]),
            "employed_post_secondary_iek":      to_int(row[5]),
            "employed_lyceum":                  to_int(row[6]),
            "employed_vocational_gymnasium":    to_int(row[7]),
            "employed_primary":                 to_int(row[8]),
            "employed_below_primary":           to_int(row[9]) if len(row) > 9 else None,
        })
    return [r for r in records if r["name_el"]]


def parse_SAM04_B14(ws):
    """Άνεργοι κατά επίπεδο εκπαίδευσης — Δήμοι."""
    rows = list(ws.iter_rows(values_only=True))
    start = find_data_start(rows)
    if start is None:
        return []

    records = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        records.append({
            "geo_level":                         to_int(row[0]),
            "geo_code":                          clean(row[1]),
            "name_el":                           clean(row[2]),
            "unemployed_total":                  to_int(row[3]),
            "unemployed_phd_masters_university": to_int(row[4]),
            "unemployed_post_secondary_iek":     to_int(row[5]),
            "unemployed_lyceum":                 to_int(row[6]),
            "unemployed_vocational_gymnasium":   to_int(row[7]),
            "unemployed_primary":                to_int(row[8]),
            "unemployed_below_primary":          to_int(row[9]) if len(row) > 9 else None,
        })
    return [r for r in records if r["name_el"]]


# ══════════════════════════════════════════════════════════
#  PARSERS — SBR01 (Enterprises by NACE & Region)
# ══════════════════════════════════════════════════════════

def _parse_SBR01(ws):
    """
    Κοινός parser για T5 (NACE section) & T7 (NACE 2-digit).
    Δομή: data ξεκινά όταν col[3] είναι αριθμός (num_legal_units).
    """
    rows = list(ws.iter_rows(values_only=True))

    start = None
    for i, row in enumerate(rows):
        if len(row) > 3 and isinstance(row[3], (int, float)) and row[0]:
            start = i
            break
    if start is None:
        return []

    records = []
    current_region = None
    for row in rows[start:]:
        region  = clean(row[0])
        if region:
            current_region = region

        nace_code  = clean(row[1])
        nace_desc  = clean(row[2])
        num_units  = to_int(row[3])
        turnover   = to_float(row[4])
        employment = to_int(row[5]) if len(row) > 5 else None

        if not current_region:
            continue
        # Παράλειψη aggregate rows (κενός κωδικός NACE = σύνολο περιφέρειας)
        if not nace_code:
            continue

        records.append({
            "region":                current_region,
            "nace_code":             nace_code,
            "nace_description":      nace_desc,
            "num_legal_units":       num_units,
            "turnover_thousand_eur": round(turnover, 2) if turnover is not None else None,
            "total_employment":      employment,
        })

    return records


def parse_SBR01_T5(ws):
    """Επιχειρήσεις ανά Section NACE Rev.2 & Περιφέρεια."""
    return _parse_SBR01(ws)


def parse_SBR01_T7(ws):
    """Επιχειρήσεις ανά 2-ψήφιο κλάδο NACE Rev.2 & Περιφέρεια."""
    return _parse_SBR01(ws)


# ══════════════════════════════════════════════════════════
#  PARSER — SEL45 (GDP/GVA wide format)
# ══════════════════════════════════════════════════════════

def parse_SEL45(ws):
    """
    ΑΕΠ/ΑΠΑ ανά κλάδο (A10) & περιφέρεια — wide format.
    Rows: περιοχές  |  Cols: κλάδος × χρονιά
    """
    rows = list(ws.iter_rows(values_only=True))

    # Βρίσκουμε τη γραμμή με χρονιές (≥10 έτη)
    year_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).strip() for v in row if v is not None]
        if sum(1 for v in vals if re.match(r"^(19|20)\d{2}\*?$", v)) >= 10:
            year_row_idx = i
            break
    if year_row_idx is None:
        return []

    # Βρίσκουμε τη γραμμή κλάδων (περιέχει "Γεωργ" ή "Agriculture" ή "Κατασκ")
    sector_row_idx = None
    for i in range(max(0, year_row_idx - 10), year_row_idx):
        row = rows[i]
        if any(v and any(kw in str(v) for kw in ["Γεωργ", "Agriculture", "Κατασκ"]) for v in row):
            sector_row_idx = i
            break
    if sector_row_idx is None:
        return []

    year_row   = rows[year_row_idx]
    sector_row = rows[sector_row_idx]

    # Forward-fill sector names (merged cells)
    SKIP_KEYWORDS = {"Κλάδοι", "Industries", "Περιφέρ", "Regions"}
    sectors_ff = []
    last_sector = None
    for v in sector_row:
        s = clean(v)
        if s and not re.match(r"^(19|20)\d{2}", s) and not any(kw in s for kw in SKIP_KEYWORDS):
            last_sector = s
        sectors_ff.append(last_sector)

    # col_index → (sector, year)
    col_map = {}
    for ci in range(1, len(year_row)):
        yr_raw = clean(year_row[ci])
        sec = sectors_ff[ci] if ci < len(sectors_ff) else None
        if yr_raw and sec:
            try:
                yr = int(str(yr_raw).replace("*", "").strip())
                col_map[ci] = (sec, yr)
            except ValueError:
                pass

    if not col_map:
        return []

    records = []
    for row in rows[year_row_idx + 1:]:
        region = clean(row[0])
        if not region:
            continue
        # Αγνόησε αγγλικές μεταφράσεις (κεφαλαία χωρίς ελληνικούς χαρακτήρες)
        if re.match(r"^[A-Z\s\-/,]+$", region):
            continue

        for ci, (sector, year) in col_map.items():
            if ci >= len(row):
                continue
            val = to_float(row[ci])
            if val is None:
                continue
            records.append({
                "region":                       region,
                "year":                         year,
                "economic_sector_A10":          sector,
                "gross_value_added_million_eur": round(val, 4),
            })

    return records


# ══════════════════════════════════════════════════════════
#  PARSER — SJO02 (Employment/Unemployment time series)
# ══════════════════════════════════════════════════════════

GREEK_MONTHS = {
    "Ιανουάριος": 1, "Φεβρουάριος": 2, "Μάρτιος": 3,  "Απρίλιος": 4,
    "Μάιος": 5,      "Ιούνιος": 6,     "Ιούλιος": 7,   "Αύγουστος": 8,
    "Σεπτέμβριος": 9,"Οκτώβριος": 10,  "Νοέμβριος": 11,"Δεκέμβριος": 12,
}


def parse_SJO02(ws):
    """
    Κατάσταση απασχόλησης — μηνιαία χρονοσειρά.
    Δομή: Έτος ως header row, μετά 12 rows με μηνιαία δεδομένα.
    Cols: month | employed | unemployed | outside_labour | unemp_rate |
          employed_sa | unemployed_sa | outside_sa | unemp_rate_sa
    """
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_year = None

    for row in rows:
        if not row or row[0] is None:
            continue

        # Έτος-header row: col[0] είναι int στη ζώνη 1990-2030
        if isinstance(row[0], (int, float)):
            yr = int(row[0])
            if 1990 <= yr <= 2030:
                current_year = yr
            continue

        # Month row: col[0] είναι string (ελληνικό μήνα)
        month_str = clean(row[0])
        if not month_str or current_year is None:
            continue
        month_num = GREEK_MONTHS.get(month_str)
        if month_num is None:
            continue

        records.append({
            "year":  current_year,
            "month": month_num,
            "month_name_el": month_str,
            # Μη εποχικά διορθωμένα
            "employed_thousands":            to_float(row[1]) if len(row) > 1 else None,
            "unemployed_thousands":          to_float(row[2]) if len(row) > 2 else None,
            "outside_labour_force_thousands": to_float(row[3]) if len(row) > 3 else None,
            "unemployment_rate_pct":         to_float(row[4]) if len(row) > 4 else None,
            # Εποχικά προσαρμοσμένα
            "employed_sa_thousands":            to_float(row[5]) if len(row) > 5 else None,
            "unemployed_sa_thousands":          to_float(row[6]) if len(row) > 6 else None,
            "outside_labour_force_sa_thousands": to_float(row[7]) if len(row) > 7 else None,
            "unemployment_rate_sa_pct":          to_float(row[8]) if len(row) > 8 else None,
        })

    return records


# ══════════════════════════════════════════════════════════
#  Routing: filename → (table_key, parser_fn)
# ══════════════════════════════════════════════════════════

def get_parser(filename, folder_name):
    """
    Επιστρέφει (table_key, parser_function) για το δεδομένο αρχείο.
    Χρησιμοποιεί συνδυασμό ονόματος αρχείου + φακέλου για disambiguation.
    """
    f      = filename.upper()
    folder = folder_name.upper()

    if "SJO02" in f:
        return "employment_unemployment", parse_SJO02
    if "SEL45" in f:
        return "gdp_gva", parse_SEL45

    # Εξαγωγή κωδικού πίνακα (A01, B03, κτλ.)
    m = re.search(r'_(A\d+|B\d+)_', f)
    if m:
        code = m.group(1)

        # SAM03 (Regions)
        if "SAM03" in f:
            parsers = {"A01": parse_SAM03_A01, "A02": parse_SAM03_A02,
                       "A04": parse_SAM03_A04, "A06": parse_SAM03_A06,
                       "A07": parse_SAM03_A07}
            return f"sam03_{code.lower()}", parsers.get(code)

        # SAM04 (Municipality — Education)
        if "SAM04" in f:
            parsers = {"B03": parse_SAM04_B03, "B09": parse_SAM04_B09,
                       "B14": parse_SAM04_B14}
            return f"sam04_{code.lower()}", parsers.get(code)

        # SAM05 (Municipality — Buildings)
        if "SAM05" in f:
            parsers = {"B01": parse_SAM05_B01, "B02": parse_SAM05_B02,
                       "B03": parse_SAM05_B03}
            return f"sam05_{code.lower()}", parsers.get(code)

    # SBR01 (Enterprises) — κωδικός πίνακα από αριθμό _05_ / _07_
    m = re.search(r'_(\d{2})_F_', f)
    if m and "SBR01" in f:
        num = m.group(1).lstrip("0")
        if num == "5":
            return "enterprises_nace_section", parse_SBR01_T5
        if num == "7":
            return "enterprises_nace_2digit", parse_SBR01_T7

    return "unknown", None


# ══════════════════════════════════════════════════════════
#  Municipality-centric aggregation
# ══════════════════════════════════════════════════════════

# Χαρτογράφηση table_key → (target_entity, section_name)
# target_entity: "municipalities" | "regions" | "national"
TABLE_ROUTING = {
    # SAM03 — Region level
    "sam03_a01": ("regions", "population_by_sex_age"),
    "sam03_a02": ("regions", "population_by_citizenship"),
    "sam03_a04": ("regions", "population_by_birth_place"),
    "sam03_a06": ("regions", "households_by_size"),
    "sam03_a07": ("regions", "one_person_households"),
    # SAM05 — Municipality level (Buildings)
    "sam05_b01": ("municipalities", "dwellings_by_occupancy_density"),
    "sam05_b02": ("municipalities", "dwellings_by_building_type"),
    "sam05_b03": ("municipalities", "households_by_size"),
    # SAM04 — Municipality level (Education)
    "sam04_b03": ("municipalities", "population_by_education_employment"),
    "sam04_b09": ("municipalities", "employed_by_education"),
    "sam04_b14": ("municipalities", "unemployed_by_education"),
    # National / Regional aggregates
    "gdp_gva":                   ("national", "gdp_gva_by_region"),
    "enterprises_nace_section":  ("national", "enterprises_by_nace_section"),
    "enterprises_nace_2digit":   ("national", "enterprises_by_nace_2digit"),
    "employment_unemployment":   ("national", "employment_unemployment"),
}


def build_output(all_data):
    """
    Οργανώνει όλα τα parsed records σε municipality-centric δομή.

    Για geo-structured δεδομένα (SAM03/04/05):
      - geo_level 5 (Δήμοι) → municipalities[geo_code][section]
      - geo_level 3/4 (Περιφέρειες/ΠΕ) → regions[geo_code][section]
      - geo_level 0/1 (Εθνικό) → παραλείπεται (πλεονάζει)

    Για region-level δεδομένα (SBR01, SEL45):
      → national[section] (indexed by region name)

    Για time-series (SJO02):
      → national[section]
    """
    municipalities = {}
    regions        = {}
    national       = {}

    def upsert(store, key, section, record):
        if key not in store:
            store[key] = {"geo_code": key}
        store[key].setdefault(section, []).append(record)

    for entry in all_data:
        tk      = entry["table_key"]
        records = entry["records"]
        if not records:
            continue

        routing = TABLE_ROUTING.get(tk)
        if routing is None:
            continue

        target, section = routing

        if target == "national":
            national.setdefault(section, []).extend(records)

        elif target in ("municipalities", "regions"):
            for rec in records:
                gl = rec.get("geo_level")
                gc = rec.get("geo_code")
                if gl is None or gc is None:
                    continue

                # Ανάθεση σε δήμο ή περιφέρεια βάσει geo_level
                if target == "municipalities":
                    # Επίπεδο 5 = Δήμος, 4 = ΠΕ, 3 = Περιφέρεια, 1/0 = ανώτερα
                    if gl == 5:
                        # Ψάχνουμε το όνομα από οποιοδήποτε πεδίο μοιάζει με γεωγραφικό όνομα
                        for candidate in (
                            rec.get("name_el"),
                            rec.get("occupancy_density"),
                            rec.get("building_type"),
                            rec.get("household_size"),
                            rec.get("education_level"),
                        ):
                            if candidate and isinstance(candidate, str) and any(
                                kw in candidate.upper() for kw in
                                ["ΔΗΜΟΣ", "ΚΟΙΝΟΤΗΤΑ", "ΠΕΡΙΦΕΡ"]
                            ):
                                entity_name = candidate
                                break
                        else:
                            entity_name = rec.get("name_el") or gc

                        if gc not in municipalities:
                            municipalities[gc] = {
                                "geo_code":  gc,
                                "geo_level": gl,
                                "name_el":   entity_name,
                            }
                        elif municipalities[gc].get("name_el") == gc:
                            municipalities[gc]["name_el"] = entity_name
                        upsert(municipalities, gc, section, rec)
                    elif gl in (3, 4):
                        name = rec.get("name_el", gc)
                        if "name_el" not in regions.get(gc, {}):
                            regions.setdefault(gc, {})["geo_code"]  = gc
                            regions[gc]["geo_level"] = gl
                            regions[gc]["name_el"]   = name
                        upsert(regions, gc, section, rec)
                    # geo_level 0/1 = εθνικό/μείζονα περιοχή — παραλείπεται

                else:  # target == "regions" (SAM03)
                    if gl in (3, 4):
                        if "name_el" not in regions.get(gc, {}):
                            regions.setdefault(gc, {})["geo_code"]  = gc
                            regions[gc]["geo_level"] = gl
                            regions[gc]["name_el"]   = rec.get("name_el", gc)
                        upsert(regions, gc, section, rec)
                    # geo_level 0/1 → παραλείπεται (aggregate rows)

    return municipalities, regions, national


# ══════════════════════════════════════════════════════════
#  Process folder
# ══════════════════════════════════════════════════════════

def process_folder(folder_name, folder_path):
    all_data = []
    tmp_dirs = []

    for fname in excel_files_in(folder_path):
        fpath = os.path.join(folder_path, fname)
        print(f"  📄 {fname}")

        actual_path = fpath

        # .xls → μετατροπή σε .xlsx
        if fname.lower().endswith(".xls"):
            converted = xls_to_xlsx(fpath)
            if converted is None:
                print("     ⚠️  Αδυναμία μετατροπής .xls — παράλειψη")
                continue
            actual_path = converted
            tmp_dirs.append(os.path.dirname(converted))
            print(f"     🔄 Μετατράπηκε σε xlsx")

        table_key, parser_fn = get_parser(fname, folder_name)

        if parser_fn is None:
            print(f"     ⚠️  Δεν βρέθηκε parser (table_key={table_key})")
            continue

        try:
            wb = openpyxl.load_workbook(actual_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"     ⚠️  openpyxl error: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                records = parser_fn(ws)
            except Exception as e:
                print(f"     ⚠️  Parser error ({sheet_name}): {e}")
                records = []

            if records:
                all_data.append({"table_key": table_key, "records": records})
                print(f"     ✅ {sheet_name}: {len(records):,} εγγραφές  [{table_key}]")
            else:
                print(f"     ⚠️  {sheet_name}: 0 εγγραφές")

    # Καθαρισμός temp dirs
    for d in tmp_dirs:
        shutil.rmtree(d, ignore_errors=True)

    return all_data


# ══════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="ΕΛΣΤΑΤ Excel → JSON (SMARTPOLIS v3)")
    ap.add_argument("--root",   required=True,  help="Root folder with ΕΛΣΤΑΤ subfolders")
    ap.add_argument("--output", default="smartpolis.json", help="Output JSON path")
    ap.add_argument("--indent", type=int, default=2)
    args = ap.parse_args()

    if not os.path.isdir(args.root):
        print(f"❌ Directory not found: {args.root}")
        return

    subfolders = sorted(
        d for d in os.listdir(args.root)
        if os.path.isdir(os.path.join(args.root, d)) and not d.startswith(".")
    )

    print(f"\n🔍 {len(subfolders)} φάκελοι στο: {args.root}\n")

    all_data = []
    folder_fields = {}

    for folder_name in subfolders:
        folder_path = os.path.join(args.root, folder_name)
        excels = excel_files_in(folder_path)
        if not excels:
            print(f"⏭️  Παράλειψη '{folder_name}' (χωρίς Excel)\n")
            continue

        print(f"📁 {folder_name}")
        folder_data = process_folder(folder_name, folder_path)
        all_data.extend(folder_data)
        folder_fields[folder_name] = load_fields(folder_path)
        total = sum(len(e["records"]) for e in folder_data)
        print(f"   → {total:,} εγγραφές σύνολο\n")

    print("🔧 Οργάνωση δεδομένων ανά δήμο / περιφέρεια...")
    municipalities, regions, national = build_output(all_data)

    output = {
        "metadata": {
            "project":      "SMARTPOLIS",
            "source":       "ΕΛΣΤΑΤ — Απογραφή 2021, ΕΕΔ, ΜΕΕ, Μητρώο Επιχειρήσεων",
            "description":  "Municipal socioeconomic data — municipality-centric structure",
            "generated_by": "xlsx2json.py v3",
            "fields_per_folder": folder_fields,
            "structure": {
                "municipalities": "Δεδομένα ανά δήμο (geo_level=5), key=geo_code",
                "regions":        "Δεδομένα ανά περιφέρεια/ΠΕ (geo_level=3,4), key=geo_code",
                "national":       "Εθνικές χρονοσειρές & περιφερειακά aggregates",
            },
            "sections": {
                "municipalities": [
                    "dwellings_by_occupancy_density",
                    "dwellings_by_building_type",
                    "households_by_size",
                    "population_by_education_employment",
                    "employed_by_education",
                    "unemployed_by_education",
                ],
                "regions": [
                    "population_by_sex_age",
                    "population_by_citizenship",
                    "population_by_birth_place",
                    "households_by_size",
                    "one_person_households",
                ],
                "national": [
                    "employment_unemployment",
                    "gdp_gva_by_region",
                    "enterprises_by_nace_section",
                    "enterprises_by_nace_2digit",
                ],
            },
        },
        "data": {
            "municipalities": municipalities,
            "regions":        regions,
            "national":       national,
        },
    }

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=args.indent, default=str)

    # Stats
    total_records = sum(len(e["records"]) for e in all_data)
    mun_sections  = sum(len([k for k in v if k not in ("geo_code","geo_level","name_el")]) for v in municipalities.values())

    print(f"\n✅  {args.output}")
    print(f"   Δήμοι:           {len(municipalities):>6,}")
    print(f"   Περιφέρειες/ΠΕ: {len(regions):>6,}")
    print(f"   Εθνικά datasets: {len(national):>6,}")
    print(f"   Σύνολο εγγραφών: {total_records:>6,}")


if __name__ == "__main__":
    main()